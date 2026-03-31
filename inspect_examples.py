#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook(r'public\QUAL TM 0027 - 01 PPAP Package.xlsx', data_only=True)

# Process Flow Example
print("="*80)
print("PROCESS FLOW EXAMPLE SHEET (5 - Process Flow - Example)")
print("="*80)
ws = wb['5 - Process Flow - Example']
print(f"Dimensions: {ws.max_row} rows x {ws.max_column} cols\n")

# Find first row with actual data
print("Looking for data rows...")
for row in range(1, 30):
    row_has_data = False
    row_str = f"Row {row:2d}: "
    for col in range(1, 11):
        val = ws.cell(row, col).value
        if val:
            row_has_data = True
            row_str += f"{ws.cell(row, col).coordinate}={str(val)[:20]} | "
    if row_has_data:
        print(row_str)

# Control Plan Example  
print("\n" + "="*80)
print("CONTROL PLAN EXAMPLE SHEET (7 - Control Plan Example)")
print("="*80)
ws2 = wb['7 - Control Plan Example']
print(f"Dimensions: {ws2.max_row} rows x {ws2.max_column} cols\n")

print("Looking for data rows...")
for row in range(1, 30):
    row_has_data = False
    row_str = f"Row {row:2d}: "
    for col in range(1, 14):
        val = ws2.cell(row, col).value
        if val:
            row_has_data = True
            row_str += f"{ws2.cell(row, col).coordinate}={str(val)[:20]} | "
    if row_has_data:
        print(row_str)

wb.close()

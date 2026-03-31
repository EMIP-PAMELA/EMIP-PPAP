import openpyxl

# Load workbook
wb = openpyxl.load_workbook(r'public/QUAL TM 0027 - 01 PPAP Package.xlsx')

print("=" * 80)
print("PROCESS FLOW DIAGRAM SHEET INSPECTION")
print("=" * 80)

ws = wb['5-Proces Flow Diagram']
print(f"\nSheet: {ws.title}")
print(f"Dimensions: {ws.max_row} rows x {ws.max_column} cols")

print("\n--- Header Area (Rows 1-15, Columns A-J) ---")
for i in range(1, 16):
    row_data = []
    for j in range(1, 11):
        cell = ws.cell(i, j)
        value = str(cell.value) if cell.value else ""
        if value:
            row_data.append(f"{cell.coordinate}:{value[:30]}")
    if row_data:
        print(f"Row {i}: {', '.join(row_data)}")

print("\n--- Looking for column headers (around row 4-6) ---")
for i in range(4, 8):
    print(f"\nRow {i}:")
    for j in range(1, 11):
        cell = ws.cell(i, j)
        if cell.value:
            print(f"  {cell.coordinate}: {cell.value}")

print("\n" + "=" * 80)
print("CONTROL PLAN SHEET INSPECTION")
print("=" * 80)

ws2 = wb['7_Process Control Plan - Form']
print(f"\nSheet: {ws2.title}")
print(f"Dimensions: {ws2.max_row} rows x {ws2.max_column} cols")

print("\n--- Header Area (Rows 1-15, Columns A-M) ---")
for i in range(1, 16):
    row_data = []
    for j in range(1, 14):
        cell = ws2.cell(i, j)
        value = str(cell.value) if cell.value else ""
        if value:
            row_data.append(f"{cell.coordinate}:{value[:30]}")
    if row_data:
        print(f"Row {i}: {', '.join(row_data)}")

print("\n--- Looking for column headers (around row 6-10) ---")
for i in range(6, 12):
    print(f"\nRow {i}:")
    for j in range(1, 14):
        cell = ws2.cell(i, j)
        if cell.value:
            print(f"  {cell.coordinate}: {cell.value}")

wb.close()

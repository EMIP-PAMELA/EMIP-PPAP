#!/usr/bin/env python3
"""
Workbook Mapping Validation Script
Inspects PPAP Package workbook to determine exact cell locations for export mapping
"""
import openpyxl
import sys

def inspect_sheet(ws, sheet_name, max_header_row=15, max_data_col=15):
    """Inspect a worksheet and print structure details"""
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} cols")
    
    # Find cells with "Part" in them (likely part number field)
    print(f"\n--- Cells containing 'Part' (likely part number location) ---")
    for row in range(1, max_header_row + 1):
        for col in range(1, max_data_col + 1):
            cell = ws.cell(row, col)
            if cell.value and 'part' in str(cell.value).lower():
                print(f"  {cell.coordinate}: {cell.value}")
    
    # Print header area
    print(f"\n--- Header Area (Rows 1-{max_header_row}) ---")
    for row in range(1, max_header_row + 1):
        cells_with_values = []
        for col in range(1, max_data_col + 1):
            cell = ws.cell(row, col)
            if cell.value:
                value_str = str(cell.value)[:40]  # Truncate long values
                cells_with_values.append(f"{cell.coordinate}='{value_str}'")
        if cells_with_values:
            print(f"Row {row:2d}: {' | '.join(cells_with_values)}")
    
    # Look for column headers (typically bold or in a specific row)
    print(f"\n--- Potential Column Header Rows (looking for step/process/operation keywords) ---")
    keywords = ['step', 'process', 'operation', 'machine', 'tool', 'characteristic', 'method', 'sample']
    for row in range(1, max_header_row + 1):
        row_values = []
        has_keyword = False
        for col in range(1, max_data_col + 1):
            cell = ws.cell(row, col)
            if cell.value:
                value_str = str(cell.value).lower()
                row_values.append(f"{cell.coordinate}='{cell.value}'")
                if any(kw in value_str for kw in keywords):
                    has_keyword = True
        if has_keyword:
            print(f"\nRow {row} (likely column headers):")
            for val in row_values:
                print(f"  {val}")

def main():
    try:
        wb = openpyxl.load_workbook(r'public\QUAL TM 0027 - 01 PPAP Package.xlsx', data_only=True)
        
        # Inspect Process Flow sheet
        if '5-Proces Flow Diagram' in wb.sheetnames:
            ws_pf = wb['5-Proces Flow Diagram']
            inspect_sheet(ws_pf, '5-Proces Flow Diagram', max_header_row=20, max_data_col=15)
        else:
            print("ERROR: Process Flow sheet not found")
            return 1
        
        # Inspect Control Plan sheet
        if '7_Process Control Plan - Form' in wb.sheetnames:
            ws_cp = wb['7_Process Control Plan - Form']
            inspect_sheet(ws_cp, '7_Process Control Plan - Form', max_header_row=20, max_data_col=15)
        else:
            print("ERROR: Control Plan sheet not found")
            return 1
        
        wb.close()
        return 0
        
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    sys.exit(main())
